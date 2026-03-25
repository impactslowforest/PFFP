"""
PFFP Data Import Script
Reads Excel file and imports data into PocketBase collections.
Clears existing data first, then imports in FK order.
"""
import sys, io, json, time
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import requests
from openpyxl import load_workbook

PB_URL = "http://127.0.0.1:8090"
EXCEL_FILE = r"c:\Users\User\OneDrive - Slow Forest\Apps\PFFP\Cloude_PFFP_16Feb2026\Xu ly_PFFP_Main data_15Feb2026.xlsx"

# Excel header → PB field name mapping (only for headers that differ)
FIELD_MAP = {
    'Farmers': {
        'Participation Year': 'participation_year',
        'Year of support': 'year_of_support',
        'ID card': 'id_card',
        'Socioeconomic Status': 'socioeconomic_status',
        'Household Circumstances': 'household_circumstances',
        'Number of coffee farm plots': 'number_of_coffee_farm_plots',
        'Supported by': 'supported_by',
        'Manage by': 'manage_by',
        'Supported Types': 'supported_types',
        'Number Farm registered for support from': 'number_farm_registered_for_support_from',
        'Total Area registered': 'total_area_registered',
        'Staff input': 'staff_input',
        # Add mappings for CamelCase headers if any
        'Farmer_ID': 'farmer_id',
        'Full_Name': 'full_name',
        'Year_Of_Birth': 'year_of_birth',
        'Gender': 'gender',
        'Phone_Number': 'phone_number',
        'Farmer_Group_Name': 'farmer_group_name',
        'Cooperative_Name': 'cooperative_name',
        'Village_Name': 'village_name',
        'Commune_Name': 'commune_name',
        'Address': 'address',
        'Ethnicity': 'ethnicity',
        'Num_Household_Members': 'num_household_members',
        'Num_Working_Members': 'num_working_members',
        'Total_Coffee_Area': 'total_coffee_area',
        'Status': 'status',
        'Activity': 'activity',
    },
    'Plots': {
        'Plot_Id': 'plot_id',
        'Farmer_ID': 'farmer_id',
        'Plot_Name': 'plot_name',
        'Area (ha)': 'area_ha',
        'Location': 'location',
        'Land use rights certificate?': 'land_use_rights_certificate',
        'Border_Natural_Forest': 'border_natural_forest',
        'Place name': 'place_name',
        'Num_Shade_Trees_Before': 'num_shade_trees_before',
        'Name_Shade_Trees_Before': 'name_shade_trees_before',
        'Num_Coffee_Trees': 'num_coffee_trees',
        'Coffee_Planted_Year': 'coffee_planted_year',
        'Notes for details (Optional)': 'notes_for_details',
        'Map Sheet': 'map_sheet',
        'Sub-mapsheet': 'sub_mapsheet',
        'Receive seedlings from': 'receive_seedlings_from',
        'Farm registered for support from': 'farm_registered_for_support_from',
        'Number of shade trees': 'number_of_shade_trees',
        'Number of shade tree species': 'number_of_shade_tree_species',
        'Status': 'status',
        'Activity': 'activity',
    },
    'Yearly_Data': {
        'Record_Id': 'record_id',
        'Farmer_ID': 'farmer_id',
        'Year': 'year',
        'Annual_Volume_Cherry': 'annual_volume_cherry',
        'Volume_High_Quality': 'volume_high_quality',
        'Total_Coffee_Income': 'total_coffee_income',
        'Fertilizers_Applied': 'fertilizers_applied',
        'Name of fertilizer': 'name_of_fertilizer',
        'Fertilizer volume': 'fertilizer_volume',
        'Fertilizer cost': 'fertilizer_cost',
        'Pesticides_Applied': 'pesticides_applied',
        'Name of Pesticides': 'name_of_pesticides',
        'Pesticides volume': 'pesticides_volume',
        'Pesticides cost': 'pesticides_cost',
        'Herbicides_Applied': 'herbicides_applied',
        'Name of Herbicides': 'name_of_herbicides',
        'Herbicides volume': 'herbicides_volume',
        'Herbicides cost': 'herbicides_cost',
        'Hired_Labor_Costs': 'hired_labor_costs',
        'Other_Costs': 'other_costs',
        'Shade_Trees_supported by': 'shade_trees_supported_by',
        'Number_Shade_Trees_Planted': 'number_shade_trees_planted',
        'Shade_Trees_Species': 'shade_trees_species',
        'Year planted': 'year_planted',
        'Shade_Trees_Died': 'shade_trees_died',
        'Survival Rate': 'survival_rate',
        'Fertiliser supported by WWF': 'fertiliser_by_wwf',
        'Lime supported by Slow': 'lime_from_slow',
        'Cover crop supported by Slow (yes/no)': 'cover_crop_from_slow',
        'Attending training capacity organized by PFFP': 'attending_training',
        'Op6_Activities': 'op6_activities',
        'Cherry sales registered to Slow': 'cherry_sales_registered_to_slow',
        'Cherry sales supplied to Slow': 'cherry_sales_supplied_to_slow',
        'Revenue from cherry sales to Slow (VND)': 'revenue_from_cherry_sales_to_slow',
        'Cherry bought by Slow via processor': 'cherry_slow_thru_processor',
    },
    'Supported': {
        'Record_Id': 'support_id',
        'Farmer ID': 'farmer_id',
        'Support code': 'support_code',
        'A live': 'a_live',
        'Supported by': 'supported_by',
        'Suppoted year': 'supported_year',
    }
}

# Collections to process in order (FK dependency)
COLLECTIONS = [
    ('Farmers',     'farmers',      'farmer_id'),
    ('Plots',       'plots',        'plot_id'),
    ('Yearly_Data', 'yearly_data',  'record_id'),
    ('Supported',   'support',      'support_id'), # Sheet='Supported' -> Collection='support'
]

def delete_all_records(collection, headers):
    """Delete all records from a collection (paginated)."""
    total_deleted = 0
    while True:
        resp = requests.get(f"{PB_URL}/api/collections/{collection}/records?perPage=200&fields=id", headers=headers)
        data = resp.json()
        items = data.get('items', [])
        if not items:
            break
        for item in items:
            r = requests.delete(f"{PB_URL}/api/collections/{collection}/records/{item['id']}", headers=headers)
            if r.status_code in (200, 204):
                total_deleted += 1
            else:
                print(f"  WARN: Failed to delete {item['id']}: {r.status_code}")
    return total_deleted

import re

def to_pb_field(header, field_map):
    """Normalize header to PocketBase field name."""
    if header in field_map:
        return field_map[header]
    # Automatic conversion: CamelCase/Spaces -> snake_case
    s = str(header).strip()
    # Replace non-alphanumeric with underscore
    s = re.sub(r'[^a-zA-Z0-9]', '_', s)
    # lowercase
    s = s.lower()
    # Collapse multiple underscores
    s = re.sub(r'_+', '_', s)
    return s.strip('_')

def read_sheet(wb, sheet_name, field_map):
    """Read a sheet and return list of dicts with PB field names."""
    if sheet_name not in wb.sheetnames:
        print(f"  WARN: Sheet '{sheet_name}' not found in Excel.")
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, values_only=False))
    if not rows:
        return []

    headers = [cell.value for cell in rows[0]]
    records = []

    for row in rows[1:]:
        values = [cell.value for cell in row]
        record = {}
        for header, value in zip(headers, values):
            if header is None:
                continue
            # Map Excel header to PB field name
            pb_field = to_pb_field(header, field_map)
            # Convert value to string for text fields, keep numbers
            if value is None:
                record[pb_field] = ""
            elif isinstance(value, (int, float)):
                # Keep numeric for number fields
                record[pb_field] = value
            else:
                record[pb_field] = str(value).strip()
        records.append(record)

    return records

def import_records(collection, records, pk_field):
    """Import records into a PocketBase collection."""
    success = 0
    errors = 0
    for i, record in enumerate(records):
        try:
            resp = requests.post(
                f"{PB_URL}/api/collections/{collection}/records",
                json=record,
                headers=HEADERS
            )
            if resp.status_code in (200, 201):
                success += 1
            else:
                errors += 1
                if errors <= 3:  # Show first 3 errors
                    print(f"  ERR [{resp.status_code}]: {record.get(pk_field, '?')} - {resp.text[:200]}")
        except Exception as e:
            errors += 1
            if errors <= 3:
                print(f"  EXCEPTION: {e}")

        # Progress
        if (i + 1) % 500 == 0:
            print(f"  ... {i+1}/{len(records)} (ok={success}, err={errors})")

    return success, errors

def main():
    print("=" * 60)
    print("PFFP Data Import")
    print("=" * 60)

    # Auth
    print("Authenticating...")
    auth_resp = requests.post(f"{PB_URL}/api/collections/_superusers/auth-with-password", json={
        "identity": "impact@slowforest.vn",
        "password": "Huongphung@452#"
    })
    if auth_resp.status_code != 200:
        print("Auth failed:", auth_resp.text)
        return
    
    token = auth_resp.json()["token"]
    global HEADERS
    HEADERS = {"Authorization": token, "Content-Type": "application/json"}

    # Load Excel
    print(f"\nLoading Excel: {EXCEL_FILE}")
    wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    # Step 1: Clear existing data (reverse FK order)
    print("\n--- CLEARING EXISTING DATA ---")
    for sheet_name, collection, pk in reversed(COLLECTIONS):  # Clear all 4
        print(f"Clearing {collection}...")
        deleted = delete_all_records(collection, HEADERS)
        print(f"  Deleted {deleted} records")

    # Step 2: Import data
    print("\n--- IMPORTING DATA ---")
    for sheet_name, collection, pk in COLLECTIONS:
        print(f"\nImporting {sheet_name} → {collection}")
        field_map = FIELD_MAP.get(sheet_name, {})
        records = read_sheet(wb, sheet_name, field_map)
        print(f"  Read {len(records)} records from Excel")

        success, errors = import_records(collection, records, pk)
        print(f"  Result: {success} imported, {errors} errors")

    # Step 3: Verify counts
    print("\n--- VERIFICATION ---")
    for sheet_name, collection, pk in COLLECTIONS:
        resp = requests.get(f"{PB_URL}/api/collections/{collection}/records?perPage=1")
        data = resp.json()
        print(f"  {collection}: {data.get('totalItems', '?')} records")

    wb.close()
    print("\nDone!")

if __name__ == "__main__":
    main()
